import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

pasta = r"C:\Users\vanclercio.pontes.BOOKERBRASIL\Documents\tcc-ufsc-contabilidade-rag\backend\data"

arquivos = []
for nome in sorted(os.listdir(pasta)):
    caminho = os.path.join(pasta, nome)
    if os.path.isfile(caminho):
        nome_sem_ext = os.path.splitext(nome)[0]
        arquivos.append(nome_sem_ext)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Base de Conhecimento"

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 80

header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
header_fill = PatternFill("solid", start_color="1F3864")
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

ws["A1"] = "#"
ws["B1"] = "Título da Página"
ws["A1"].font = header_font
ws["B1"].font = header_font
ws["A1"].fill = header_fill
ws["B1"].fill = header_fill
ws["A1"].alignment = center
ws["B1"].alignment = center
ws.row_dimensions[1].height = 20

alt_fill = PatternFill("solid", start_color="DCE6F1")
normal_font = Font(name="Arial", size=10)

for i, nome in enumerate(arquivos, start=1):
    row = i + 1
    ws.cell(row=row, column=1, value=i).alignment = center
    ws.cell(row=row, column=2, value=nome).alignment = left
    ws.cell(row=row, column=1).font = normal_font
    ws.cell(row=row, column=2).font = normal_font
    if i % 2 == 0:
        ws.cell(row=row, column=1).fill = alt_fill
        ws.cell(row=row, column=2).fill = alt_fill

total_row = len(arquivos) + 2
ws.cell(row=total_row, column=1, value="Total")
ws.cell(row=total_row, column=2, value=f"=COUNTA(B2:B{total_row-1})")
ws.cell(row=total_row, column=1).font = Font(bold=True, name="Arial", size=10)
ws.cell(row=total_row, column=2).font = Font(bold=True, name="Arial", size=10)

output = r"inventario_base_sispetro.xlsx"
wb.save(output)
print(f"Arquivo salvo: {output} — {len(arquivos)} páginas encontradas.")